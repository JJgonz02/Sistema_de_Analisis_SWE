# export_results.py
# Exportación de resultados del análisis SWE.
#
# Genera dos archivos en la carpeta que elija el usuario:
#   - reporte_swe.pdf  : portada con métricas globales + todas las gráficas
#   - metricas_frames.xlsx : tabla detallada de estadísticas por fotograma
#
# La función principal es exportar_resultados(resultado, carpeta_destino).
# Se llama desde AnalysisWindow al pulsar el botón "Exportar resultados".

from __future__ import annotations

import io
import datetime
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image as RLImage,
    Table, TableStyle, HRFlowable, PageBreak,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─────────────────────────────────────────────────────────────────────────────
# Paleta (coherente con estilos.py)
# ─────────────────────────────────────────────────────────────────────────────

_C_BG     = colors.HexColor("#0f1117")
_C_PANEL  = colors.HexColor("#1a1f2e")
_C_ACCENT = colors.HexColor("#378ADD")
_C_TEXT   = colors.HexColor("#c8cfe8")
_C_DIM    = colors.HexColor("#6a7290")
_C_WARN   = colors.HexColor("#FFB800")
_C_PINK   = colors.HexColor("#ff4f7b")
_C_CYAN   = colors.HexColor("#00e5ff")
_C_BORDER = colors.HexColor("#2a2f45")

# Para matplotlib
_M_BG    = "#0f1117"
_M_PANEL = "#12151f"
_M_ACCENT= "#378ADD"
_M_TEXT  = "#c8cfe8"
_M_DIM   = "#6a7290"
_M_BORDER= "#2a2f45"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers de gráficos matplotlib → bytes PNG
# ─────────────────────────────────────────────────────────────────────────────

def _fig_a_bytes(fig: plt.Figure) -> bytes:
    """Convierte una figura matplotlib en bytes PNG para incrustar en el PDF."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _figura_mapas(resultado: dict) -> bytes:
    """Mapa de rigidez media y mapa de std en una figura 1×2."""
    r   = resultado
    fig, axes = plt.subplots(1, 2, figsize=(10, 3.8), facecolor=_M_BG)
    fig.subplots_adjust(left=0.04, right=0.96, top=0.88, bottom=0.06, wspace=0.3)

    im1 = axes[0].imshow(r["mapa_medio"], cmap="jet", vmin=0, vmax=300, aspect="auto")
    cb1 = plt.colorbar(im1, ax=axes[0], shrink=0.88)
    cb1.set_label("kPa", color=_M_TEXT, fontsize=8)
    plt.setp(cb1.ax.yaxis.get_ticklabels(), color=_M_TEXT, fontsize=7)
    axes[0].set_title("Rigidez media espacial", color=_M_TEXT, fontsize=10)
    axes[0].axis("off")

    im2 = axes[1].imshow(r["mapa_std"], cmap="hot", aspect="auto")
    cb2 = plt.colorbar(im2, ax=axes[1], shrink=0.88)
    cb2.set_label("kPa", color=_M_TEXT, fontsize=8)
    plt.setp(cb2.ax.yaxis.get_ticklabels(), color=_M_TEXT, fontsize=7)
    axes[1].set_title("Variabilidad temporal (std)", color=_M_TEXT, fontsize=10)
    axes[1].axis("off")

    fig.suptitle("Mapas espaciales", color=_M_TEXT, fontsize=11)
    return _fig_a_bytes(fig)


def _figura_histograma_global(resultado: dict) -> bytes:
    """Histograma global con líneas de media, mediana, IQR y P90."""
    r   = resultado
    sg  = r["stats_globales"]
    todos = r["volumen_kpa"][~np.isnan(r["volumen_kpa"])]
    bins  = np.linspace(0, 300, 60)

    fig, ax = plt.subplots(figsize=(10, 3.5), facecolor=_M_BG)
    fig.subplots_adjust(left=0.07, right=0.97, top=0.88, bottom=0.15)

    ax.hist(todos, bins=bins, color=_M_ACCENT, edgecolor="none", alpha=0.85)
    ax.axvline(sg["media"],   color="#FFB800", lw=1.5, label=f"Media: {sg['media']:.1f} kPa")
    ax.axvline(sg["mediana"], color="#ff4f7b", lw=1.5, ls="--",
               label=f"Mediana: {sg['mediana']:.1f} kPa")
    ax.axvspan(sg["p25"], sg["p75"], alpha=0.15, color="#FFB800",
               label=f"IQR: {sg['p25']:.1f}–{sg['p75']:.1f} kPa")
    ax.axvline(sg["p90"], color="#00e5ff", lw=1.2, ls=":",
               label=f"P90: {sg['p90']:.1f} kPa")

    ax.set_facecolor(_M_PANEL)
    ax.set_xlabel("Rigidez (kPa)", color=_M_TEXT, fontsize=9)
    ax.set_ylabel("Frecuencia",    color=_M_TEXT, fontsize=9)
    ax.set_title(f"Distribución global — {r['n_frames']} fotogramas",
                 color=_M_TEXT, fontsize=10)
    ax.set_xlim(0, 300)
    ax.tick_params(colors=_M_DIM, labelsize=8)
    ax.legend(fontsize=8, facecolor=_M_PANEL, edgecolor=_M_BORDER, labelcolor=_M_TEXT)
    for sp in ax.spines.values(): sp.set_edgecolor(_M_BORDER)

    return _fig_a_bytes(fig)


def _figura_temporal(resultado: dict) -> bytes:
    """Evolución temporal (media ± std) y cobertura por fotograma."""
    r  = resultado
    sg = r["stats_globales"]
    nf = r["n_frames"]
    eje = np.arange(nf)
    mpf = r["media_por_frame"]
    spf = r["std_por_frame"]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 3.5), facecolor=_M_BG)
    fig.subplots_adjust(left=0.07, right=0.97, top=0.88, bottom=0.15, wspace=0.35)

    # Evolución media ± std
    ax1.plot(eje, mpf, color=_M_ACCENT, lw=1.3, label="Media")
    ax1.fill_between(eje, mpf - spf, mpf + spf, alpha=0.2, color=_M_ACCENT, label="±1 std")
    ax1.axhline(sg["media"], color="#FFB800", lw=1, ls="--",
                label=f"Media global: {sg['media']:.1f} kPa")
    ax1.set_facecolor(_M_PANEL)
    ax1.set_xlabel("Fotograma", color=_M_TEXT, fontsize=9)
    ax1.set_ylabel("kPa",       color=_M_TEXT, fontsize=9)
    ax1.set_title("Evolución temporal", color=_M_TEXT, fontsize=10)
    ax1.set_xlim(0, nf - 1)
    ax1.tick_params(colors=_M_DIM, labelsize=8)
    ax1.legend(fontsize=8, facecolor=_M_PANEL, edgecolor=_M_BORDER, labelcolor=_M_TEXT)
    for sp in ax1.spines.values(): sp.set_edgecolor(_M_BORDER)

    # Cobertura por frame
    ax2.plot(eje, r["cobertura_frames"], color="#00e5ff", lw=1.3)
    ax2.axhline(sg["cobertura_media"], color="#FFB800", lw=1, ls="--",
                label=f"Media: {sg['cobertura_media']:.1f}%")
    ax2.set_facecolor(_M_PANEL)
    ax2.set_xlabel("Fotograma",     color=_M_TEXT, fontsize=9)
    ax2.set_ylabel("Cobertura (%)", color=_M_TEXT, fontsize=9)
    ax2.set_title("Cobertura válida por fotograma", color=_M_TEXT, fontsize=10)
    ax2.set_xlim(0, nf - 1)
    ax2.set_ylim(0, 100)
    ax2.tick_params(colors=_M_DIM, labelsize=8)
    ax2.legend(fontsize=8, facecolor=_M_PANEL, edgecolor=_M_BORDER, labelcolor=_M_TEXT)
    for sp in ax2.spines.values(): sp.set_edgecolor(_M_BORDER)

    return _fig_a_bytes(fig)


# ─────────────────────────────────────────────────────────────────────────────
# Construcción del PDF
# ─────────────────────────────────────────────────────────────────────────────

def _estilos_pdf():
    """Devuelve un dict de estilos ReportLab personalizados para el informe."""
    base = getSampleStyleSheet()

    estilos = {
        "titulo": ParagraphStyle(
            "titulo",
            parent=base["Title"],
            fontSize=20,
            textColor=_C_TEXT,
            backColor=_C_BG,
            alignment=TA_CENTER,
            spaceAfter=4,
        ),
        "subtitulo": ParagraphStyle(
            "subtitulo",
            parent=base["Normal"],
            fontSize=11,
            textColor=_C_DIM,
            alignment=TA_CENTER,
            spaceAfter=16,
        ),
        "seccion": ParagraphStyle(
            "seccion",
            parent=base["Heading2"],
            fontSize=12,
            textColor=_C_ACCENT,
            backColor=_C_BG,
            spaceBefore=14,
            spaceAfter=6,
            alignment=TA_LEFT,
        ),
        "normal": ParagraphStyle(
            "normal",
            parent=base["Normal"],
            fontSize=9,
            textColor=_C_TEXT,
            backColor=_C_BG,
            leading=14,
        ),
        "pie": ParagraphStyle(
            "pie",
            parent=base["Normal"],
            fontSize=8,
            textColor=_C_DIM,
            alignment=TA_CENTER,
        ),
    }
    return estilos


def _tabla_metricas_globales(sg: dict, n_frames: int, estilos: dict) -> Table:
    """Construye la tabla de métricas globales para el PDF."""
    filas = [
        ["Métrica",              "Valor"],
        ["Fotogramas analizados",str(n_frames)],
        ["Media global",         f"{sg['media']:.2f} kPa"],
        ["Mediana",              f"{sg['mediana']:.2f} kPa"],
        ["Desviación estándar",  f"{sg['std']:.2f} kPa"],
        ["Percentil 25 (P25)",   f"{sg['p25']:.2f} kPa"],
        ["Percentil 75 (P75)",   f"{sg['p75']:.2f} kPa"],
        ["Percentil 90 (P90)",   f"{sg['p90']:.2f} kPa"],
        ["Mínimo",               f"{sg['minimo']:.2f} kPa"],
        ["Máximo",               f"{sg['maximo']:.2f} kPa"],
        ["Coef. variación (CV)", f"{sg['cv']:.2f}%"],
        ["Cobertura media",      f"{sg['cobertura_media']:.2f}%"],
        ["Píxeles válidos totales", f"{sg['n_total']:,}"],
    ]

    ts = TableStyle([
        # Encabezado
        ("BACKGROUND",  (0, 0), (-1, 0), _C_ACCENT),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 10),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        # Filas de datos
        ("FONTSIZE",    (0, 1), (-1, -1), 9),
        ("TEXTCOLOR",   (0, 1), (-1, -1), _C_TEXT),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_C_PANEL, _C_BG]),
        ("ALIGN",       (1, 1), (1, -1), "CENTER"),
        # Bordes
        ("GRID",        (0, 0), (-1, -1), 0.5, _C_BORDER),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",(0, 0), (-1, -1), 10),
    ])

    return Table(filas, colWidths=[9 * cm, 5 * cm], style=ts)


def generar_pdf(resultado: dict, ruta_pdf: str):
    """
    Genera el informe PDF en ruta_pdf.
    Incluye: portada con métricas, mapa espacial, histograma global
    y gráficos de evolución temporal.
    """
    sg      = resultado["stats_globales"]
    estilos = _estilos_pdf()
    fecha   = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")

    doc = SimpleDocTemplate(
        ruta_pdf,
        pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm,  bottomMargin=2 * cm,
        title="Reporte SWE — Elastografía",
    )

    # Fondo oscuro para todas las páginas
    def _fondo(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(_C_BG)
        canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
        # Franja de pie de página
        canvas.setFillColor(_C_PANEL)
        canvas.setStrokeColor(_C_BORDER)
        canvas.rect(0, 0, A4[0], 1.1 * cm, fill=1, stroke=0)
        canvas.setFillColor(_C_DIM)
        canvas.setFont("Helvetica", 7)
        canvas.drawCentredString(A4[0] / 2, 0.4 * cm,
                                 f"Análisis SWE — {fecha}  |  pág. {doc.page}")
        canvas.restoreState()

    story = []

    # ── Portada ──────────────────────────────────────────────────────────────
    story.append(Spacer(1, 1.2 * cm))
    story.append(Paragraph("Reporte de Elastografía SWE", estilos["titulo"]))
    story.append(Paragraph(f"Generado el {fecha}", estilos["subtitulo"]))
    story.append(HRFlowable(width="100%", thickness=1, color=_C_ACCENT, spaceAfter=18))

    # ── Tabla de métricas globales ────────────────────────────────────────────
    story.append(Paragraph("Métricas globales", estilos["seccion"]))
    story.append(_tabla_metricas_globales(sg, resultado["n_frames"], estilos))
    story.append(Spacer(1, 0.5 * cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_C_BORDER, spaceAfter=10))

    # ── Mapas espaciales ─────────────────────────────────────────────────────
    story.append(Paragraph("Mapas espaciales", estilos["seccion"]))
    png_mapas = _figura_mapas(resultado)
    story.append(RLImage(io.BytesIO(png_mapas), width=16 * cm, height=6 * cm))
    story.append(Paragraph(
        "Izquierda: rigidez media espacial (kPa, promedio de todos los fotogramas). "
        "Derecha: variabilidad temporal (desviación estándar entre fotogramas).",
        estilos["pie"],
    ))
    story.append(Spacer(1, 0.4 * cm))

    # ── Histograma global ─────────────────────────────────────────────────────
    story.append(Paragraph("Distribución global de rigidez", estilos["seccion"]))
    png_hist = _figura_histograma_global(resultado)
    story.append(RLImage(io.BytesIO(png_hist), width=16 * cm, height=5.6 * cm))
    story.append(Paragraph(
        "Histograma de todos los píxeles válidos del volumen completo. "
        "Se muestran media, mediana, rango intercuartílico (IQR) y percentil 90.",
        estilos["pie"],
    ))
    story.append(Spacer(1, 0.4 * cm))

    # ── Evolución temporal y cobertura ────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("Evolución temporal", estilos["seccion"]))
    png_temp = _figura_temporal(resultado)
    story.append(RLImage(io.BytesIO(png_temp), width=16 * cm, height=5.6 * cm))
    story.append(Paragraph(
        "Izquierda: media ± desviación estándar por fotograma. "
        "Derecha: porcentaje de píxeles con color SWE válido (cobertura) por fotograma.",
        estilos["pie"],
    ))

    doc.build(story, onFirstPage=_fondo, onLaterPages=_fondo)


# ─────────────────────────────────────────────────────────────────────────────
# Construcción del Excel
# ─────────────────────────────────────────────────────────────────────────────

def _estilo_celda(ws, fila: int, col: int,
                  negrita=False, fondo=None, alineacion="center",
                  fuente_color="C8CFE8"):
    """Aplica estilos visuales a una celda de openpyxl."""
    cell = ws.cell(row=fila, column=col)
    cell.font = Font(name="Arial", bold=negrita, color=fuente_color)
    cell.alignment = Alignment(horizontal=alineacion, vertical="center")
    if fondo:
        cell.fill = PatternFill("solid", start_color=fondo)
    borde = Side(style="thin", color="2A2F45")
    cell.border = Border(left=borde, right=borde, top=borde, bottom=borde)
    return cell


def generar_excel(resultado: dict, ruta_excel: str):
    """
    Genera el Excel en ruta_excel con dos hojas:
      - Resumen global : métricas globales
      - Por fotograma  : media, mediana, std, P25, P75, cobertura de cada frame
    """
    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen global ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen global"
    ws1.sheet_view.showGridLines = False

    # Encabezado
    ws1.merge_cells("A1:C1")
    c = ws1["A1"]
    c.value     = "Reporte de Elastografía SWE — Métricas Globales"
    c.font      = Font(name="Arial", bold=True, size=14, color="C8CFE8")
    c.fill      = PatternFill("solid", start_color="1A1F2E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:C2")
    fecha_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    c2 = ws1["A2"]
    c2.value     = f"Generado el {fecha_str}"
    c2.font      = Font(name="Arial", size=9, color="6A7290")
    c2.fill      = PatternFill("solid", start_color="0F1117")
    c2.alignment = Alignment(horizontal="center")

    # Cabecera de la tabla
    for col, titulo in enumerate(["Métrica", "Valor", "Unidad"], start=1):
        cell = _estilo_celda(ws1, 4, col, negrita=True, fondo="378ADD",
                             fuente_color="FFFFFF")
        cell.value = titulo

    sg      = resultado["stats_globales"]
    n_frames= resultado["n_frames"]
    filas   = [
        ("Fotogramas analizados",    n_frames,             "—"),
        ("Media global",             round(sg["media"],    2), "kPa"),
        ("Mediana",                  round(sg["mediana"],  2), "kPa"),
        ("Desviación estándar",      round(sg["std"],      2), "kPa"),
        ("Percentil 25 (P25)",       round(sg["p25"],      2), "kPa"),
        ("Percentil 75 (P75)",       round(sg["p75"],      2), "kPa"),
        ("Percentil 90 (P90)",       round(sg["p90"],      2), "kPa"),
        ("Mínimo",                   round(sg["minimo"],   2), "kPa"),
        ("Máximo",                   round(sg["maximo"],   2), "kPa"),
        ("Coeficiente de variación", round(sg["cv"],       2), "%"),
        ("Cobertura media",          round(sg["cobertura_media"], 2), "%"),
        ("Píxeles válidos totales",  sg["n_total"],             "px"),
    ]

    fondos = ["1A1F2E", "12151F"]
    for i, (metrica, valor, unidad) in enumerate(filas):
        fila = i + 5
        fondo = fondos[i % 2]
        _estilo_celda(ws1, fila, 1, fondo=fondo, alineacion="left").value  = metrica
        _estilo_celda(ws1, fila, 2, fondo=fondo, alineacion="center").value = valor
        _estilo_celda(ws1, fila, 3, fondo=fondo, alineacion="center",
                      fuente_color="6A7290").value = unidad

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 10

    # ── Hoja 2: Métricas por fotograma ───────────────────────────────────────
    ws2 = wb.create_sheet("Por fotograma")
    ws2.sheet_view.showGridLines = False

    # Encabezado
    ws2.merge_cells("A1:H1")
    c3 = ws2["A1"]
    c3.value     = "Métricas por fotograma — Elastografía SWE"
    c3.font      = Font(name="Arial", bold=True, size=13, color="C8CFE8")
    c3.fill      = PatternFill("solid", start_color="1A1F2E")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    cabeceras = [
        "Fotograma", "Media (kPa)", "Mediana (kPa)", "Std (kPa)",
        "P25 (kPa)", "P75 (kPa)", "Cobertura (%)", "Píxeles válidos",
    ]
    for col, titulo in enumerate(cabeceras, start=1):
        cell = _estilo_celda(ws2, 3, col, negrita=True, fondo="378ADD",
                             fuente_color="FFFFFF")
        cell.value = titulo

    mpf = resultado["media_por_frame"]
    spf = resultado["std_por_frame"]
    cob = resultado["cobertura_frames"]
    vol = resultado["volumen_kpa"]

    for i in range(n_frames):
        fila   = i + 4
        fondo  = fondos[i % 2]
        kpa_f  = vol[i]
        validos = kpa_f[~np.isnan(kpa_f)]
        n_val  = int(len(validos))

        if n_val > 0:
            med    = round(float(np.nanmedian(kpa_f)), 2)
            p25    = round(float(np.nanpercentile(kpa_f, 25)), 2)
            p75    = round(float(np.nanpercentile(kpa_f, 75)), 2)
        else:
            med = p25 = p75 = ""

        vals = [
            i + 1,
            round(float(mpf[i]), 2) if not np.isnan(mpf[i]) else "",
            med,
            round(float(spf[i]), 2) if not np.isnan(spf[i]) else "",
            p25,
            p75,
            round(float(cob[i]), 2),
            n_val,
        ]
        for col, valor in enumerate(vals, start=1):
            _estilo_celda(ws2, fila, col, fondo=fondo, alineacion="center").value = valor

    # Anchos de columnas
    anchos = [12, 14, 15, 12, 12, 12, 15, 16]
    for col, ancho in enumerate(anchos, start=1):
        ws2.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = ancho

    # Fijar fila de cabecera
    ws2.freeze_panes = "A4"

    wb.save(ruta_excel)


# ─────────────────────────────────────────────────────────────────────────────
# Función pública
# ─────────────────────────────────────────────────────────────────────────────

def exportar_resultados(resultado: dict, carpeta_destino: str):
    """
    Punto de entrada principal.
    Genera ambos archivos en carpeta_destino y devuelve sus rutas.

    Parámetros
    ----------
    resultado       : dict devuelto por AnalysisWorker.finished
    carpeta_destino : directorio donde se guardarán los archivos

    Retorna
    -------
    (ruta_pdf, ruta_excel) : tupla de cadenas con las rutas absolutas
    """
    import os
    ruta_pdf   = os.path.join(carpeta_destino, "reporte_swe.pdf")
    ruta_excel = os.path.join(carpeta_destino, "metricas_fotograma_a_fotograma.xlsx")

    generar_pdf(resultado,   ruta_pdf)
    generar_excel(resultado, ruta_excel)

    return ruta_pdf, ruta_excel
